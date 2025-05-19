import pandas as pd
import calendar
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Load input data
input_file = "visits.xlsx"
df = pd.read_excel(input_file)

# Ensure Date column is datetime
df['Date'] = pd.to_datetime(df['Date'])

# Extract year-month tuples
df['YearMonth'] = df['Date'].dt.to_period('M')

# Create workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

for period in sorted(df['YearMonth'].unique()):
    year = period.year
    month = period.month
    month_name = period.strftime('%B %Y')
    
    # Create sheet
    ws = wb.create_sheet(title=period.strftime('%b_%Y'))

    # Calendar setup
    cal = calendar.Calendar(firstweekday=6)  # Sunday = 6
    month_days = list(cal.itermonthdates(year, month))

    # Prepare a matrix: 6 rows (weeks) x 7 cols (Sun-Sat)
    calendar_matrix = [[] for _ in range(6)]
    week_idx = 0

    for i, day in enumerate(month_days):
        if day.month != month:
            continue  # Skip days from previous/next month
        col = day.weekday()
        if col == 6:  # Sunday
            if i != 0:
                week_idx += 1
            col = 0
        else:
            col += 1

        # Add the day and corresponding visits
        day_visits = df[df['Date'].dt.date == day]['Facility'].tolist()
        content = f"{day.day}\n" + "\n".join(day_visits)
        calendar_matrix[week_idx].append((col, content))

    # Write header
    for col_num, day_name in enumerate(['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday'], 1):
        ws.cell(row=1, column=col_num, value=day_name)

    # Write calendar
    for row_num, week in enumerate(calendar_matrix, start=2):
        for col, content in week:
            cell = ws.cell(row=row_num, column=col + 1, value=content)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Adjust column width and row height
    for col_num in range(1, 8):
        ws.column_dimensions[get_column_letter(col_num)].width = 20
    for row in range(2, 8):
        ws.row_dimensions[row].height = 100

# Save workbook
output_file = "calendar_visits.xlsx"
wb.save(output_file)
print(f"Calendar saved to {output_file}")